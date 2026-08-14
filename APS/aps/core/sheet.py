#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""表格引擎：xlsx 读写。

模型：sheets = [{name, rows: [[cell,...],...], col_widths, row_heights}]
单元格 cell = 字符串（或数字），None 表示空。
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class SheetEngine:
    def __init__(self, path: str = None):
        self.path = path
        self.sheets = []          # [{name, rows, col_widths}]
        self.active = 0
        self.dirty = False
        if path:
            self._load(path)
        else:
            self._new_default()

    def _new_default(self):
        ws = {"name": "Sheet1", "rows": [[""] * 8 for _ in range(20)],
              "col_widths": {}}
        self.sheets = [ws]
        self.dirty = True

    # ---------------- 模型 ----------------
    def sheet(self, idx=None) -> dict:
        return self.sheets[self.active if idx is None else idx]

    def cell(self, r, c, idx=None) -> str:
        ws = self.sheet(idx)
        rows = ws["rows"]
        if r < len(rows) and c < len(rows[r]):
            v = rows[r][c]
            return "" if v is None else str(v)
        return ""

    def set_cell(self, r, c, value, idx=None):
        ws = self.sheet(idx)
        rows = ws["rows"]
        while len(rows) <= r:
            rows.append([""] * max(len(rows[0]) if rows else 8, c + 1))
        if c >= len(rows[r]):
            rows[r] += [""] * (c + 1 - len(rows[r]))
        rows[r][c] = value
        self.dirty = True

    def add_sheet(self, name: str):
        self.sheets.append({"name": name, "rows": [[""] * 8 for _ in range(20)],
                            "col_widths": {}})
        self.active = len(self.sheets) - 1
        self.dirty = True

    def rename_sheet(self, idx: int, name: str):
        self.sheets[idx]["name"] = name
        self.dirty = True

    def remove_sheet(self, idx: int):
        if len(self.sheets) > 1:
            del self.sheets[idx]
            self.active = min(self.active, len(self.sheets) - 1)
            self.dirty = True

    # ---------------- 读写 ----------------
    def _load(self, path: str):
        wb = load_workbook(path, data_only=True)
        self.sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else v for v in row])
            if not rows:
                rows = [[""] * 8]
            widths = {}
            for i, dim in ws.column_dimensions.items():
                if dim.width:
                    try:
                        widths[int(i.replace("COL", "")) - 1] = int(dim.width)
                    except Exception:
                        pass
            self.sheets.append({"name": ws.title, "rows": rows, "col_widths": widths})
        self.active = 0
        self.dirty = False

    def save(self, path: str):
        wb = Workbook()
        wb.remove(wb.active)
        for ws in self.sheets:
            w = wb.create_sheet(title=ws["name"][:31])
            for r, row in enumerate(ws["rows"], 1):
                for c, v in enumerate(row, 1):
                    if v != "" and v is not None:
                        w.cell(row=r, column=c, value=v)
            for c, width in ws["col_widths"].items():
                w.column_dimensions[get_column_letter(c + 1)].width = width
        wb.save(path)
        self.path = path
        self.dirty = False

    def to_text(self) -> str:
        lines = []
        for i, ws in enumerate(self.sheets):
            lines.append(f"## 工作表：{ws['name']}")
            for row in ws["rows"]:
                vals = ["" if v is None else str(v) for v in row]
                if any(v.strip() for v in vals):
                    lines.append(" | ".join(vals))
        return "\n".join(lines)
