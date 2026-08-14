#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""表格编辑视图：网格单元格编辑 + 公式栏 + 多工作表。"""
import gi

gi.require_version("Gtk", "4.0")
from gi.repository import Gtk, GLib, Pango

from aps.core.sheet import SheetEngine
from openpyxl.utils import get_column_letter


class SheetView(Gtk.Box):
    def __init__(self, engine: SheetEngine = None):
        super().__init__(orientation=Gtk.Orientation.VERTICAL, spacing=2)
        self.engine = engine or SheetEngine(None)
        self.cur_r = 0
        self.cur_c = 0
        self._loading = False

        # 公式栏
        fbar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=4)
        self.ref_label = Gtk.Label(label="A1")
        self.ref_label.add_css_class("dim-label")
        self.ref_label.set_size_request(56, -1)
        fbar.append(self.ref_label)
        self.formula_entry = Gtk.Entry()
        self.formula_entry.set_placeholder_text("值或公式（回车确认）")
        self.formula_entry.set_hexpand(True)
        self.formula_entry.connect("activate", self._on_formula_enter)
        fbar.append(self.formula_entry)
        self.append(fbar)

        # 表格网格
        self.grid = Gtk.Grid()
        self.grid.set_row_spacing(0)
        self.grid.set_column_spacing(0)
        sw = Gtk.ScrolledWindow()
        sw.set_child(self.grid)
        sw.set_vexpand(True)
        self.append(sw)

        # 工作表标签
        self.sheet_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=4)
        self.sheet_bar.set_margin_top(4)
        self.append(self.sheet_bar)

        self._render_sheet_bar()
        self._render_grid()

    # ------------------------------------------------------------------
    def _render_sheet_bar(self):
        while (c := self.sheet_bar.get_first_child()) is not None:
            self.sheet_bar.remove(c)
        for i, ws in enumerate(self.engine.sheets):
            b = Gtk.ToggleButton(label=ws["name"])
            b.set_active(i == self.engine.active)
            b.connect("toggled", self._on_sheet_toggled, i)
            self.sheet_bar.append(b)

    def _on_sheet_toggled(self, btn, idx):
        if btn.get_active() and idx != self.engine.active:
            self._save_current_cell()
            self.engine.active = idx
            self.cur_r = self.cur_c = 0
            self._render_grid()
            self._render_sheet_bar()

    # ------------------------------------------------------------------
    def _render_grid(self):
        self._loading = True
        while (c := self.grid.get_first_child()) is not None:
            self.grid.remove(c)
        ws = self.engine.sheet()
        rows = ws["rows"]
        max_c = max(len(r) for r in rows) if rows else 8
        for r, row in enumerate(rows[:60]):
            # 行号
            lbl = Gtk.Label(label=str(r + 1))
            lbl.add_css_class("dim-label")
            lbl.set_size_request(36, -1)
            self.grid.attach(lbl, 0, r + 1, 1, 1)
            for c in range(min(max_c, 26)):
                if r == 0:
                    col = Gtk.Label(label=get_column_letter(c + 1))
                    col.add_css_class("dim-label")
                    self.grid.attach(col, c + 1, 0, 1, 1)
                entry = Gtk.Entry()
                entry.set_width_chars(10)
                v = self.engine.cell(r, c)
                entry.set_text(v if v != "" else "")
                entry.connect("changed", self._on_cell_changed, r, c)
                # GTK4：焦点事件用 EventControllerFocus（focus-in-event 是 GTK3 信号）
                ctl = Gtk.EventControllerFocus()
                ctl.connect("enter", self._on_cell_focus, r, c, entry)
                entry.add_controller(ctl)
                self.grid.attach(entry, c + 1, r + 1, 1, 1)
        self._loading = False

    def _on_cell_changed(self, entry, r, c):
        if self._loading:
            return
        self.engine.set_cell(r, c, entry.get_text())

    def _on_cell_focus(self, ctl, r, c, widget):
        self.cur_r, self.cur_c = r, c
        self.ref_label.set_text(f"{get_column_letter(c + 1)}{r + 1}")
        self.formula_entry.set_text(self.engine.cell(r, c))
        return False

    def _save_current_cell(self):
        pass

    def _on_formula_enter(self, *_):
        self.engine.set_cell(self.cur_r, self.cur_c, self.formula_entry.get_text())
        self._render_grid()

    # ------------------------------------------------------------------
    def action(self, action: str):
        if action == "ins-row":
            ws = self.engine.sheet()
            ws["rows"].insert(self.cur_r + 1, [""] * max(len(ws["rows"][0]) if ws["rows"] else 8, self.cur_c + 1))
        elif action == "del-row":
            ws = self.engine.sheet()
            if len(ws["rows"]) > 1:
                ws["rows"].pop(self.cur_r)
                self.cur_r = max(0, self.cur_r - 1)
        elif action == "ins-col":
            ws = self.engine.sheet()
            for row in ws["rows"]:
                row.insert(self.cur_c + 1, "")
        elif action == "del-col":
            ws = self.engine.sheet()
            for row in ws["rows"]:
                if self.cur_c < len(row):
                    row.pop(self.cur_c)
            self.cur_c = max(0, self.cur_c - 1)
        elif action == "new-sheet":
            self.engine.add_sheet(f"Sheet{len(self.engine.sheets) + 1}")
            self.cur_r = self.cur_c = 0
        elif action == "rename-sheet":
            self._rename_dialog()
            return
        elif action in ("sum", "avg"):
            self._compute(action)
        self.engine.dirty = True
        self._render_grid()
        self._render_sheet_bar()

    def _compute(self, op):
        total, n = 0, 0
        for c in range(self.cur_c):
            v = self.engine.cell(self.cur_r, c)
            try:
                total += float(v)
                n += 1
            except (ValueError, TypeError):
                pass
        if op == "avg":
            result = total / n if n else ""
        else:
            result = str(total)
        self.engine.set_cell(self.cur_r, self.cur_c, result)
        self.formula_entry.set_text(result)

    def _rename_dialog(self):
        dlg = Gtk.Dialog(title="重命名工作表", modal=True)
        dlg.set_default_size(300, -1)
        entry = Gtk.Entry()
        entry.set_text(self.engine.sheet()["name"])
        dlg.get_content_area().append(entry)
        ok = dlg.add_button("确定", Gtk.ResponseType.OK)
        dlg.add_button("取消", Gtk.ResponseType.CANCEL)
        ok.connect("clicked", lambda *_: self.engine.rename_sheet(self.engine.active, entry.get_text()))
        dlg.connect("response", lambda d, r: d.destroy())
        dlg.present()
